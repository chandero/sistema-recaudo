"""
Servicio para exportación de obligaciones a Excel basado en rangos de resoluciones.
Formato inspirado en el archivo de LOTE 15-3 de Alcaldía de Girón.
"""
from typing import List, Tuple
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl import utils as openpyxl_utils
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlmodel import Session, select
from app.models.obligation import Obligation
from app.models.client import Client
from app.models.tenant import Tenant


class ExcelExportService:
    """
    Servicio para exportación de datos a Excel con formato personalizado.
    """

    @staticmethod
    def export_obligations_by_resolution_range(
        db: Session,
        resolution_from: int,
        resolution_to: int
    ) -> BytesIO:
        """
        Exporta obligaciones en un rango de números de resolución a un archivo Excel con formato.
        El formato sigue el estándar de la Alcaldía de Girón.
        
        Args:
            db: Sesión de base de datos
            resolution_from: Número de resolución inicial
            resolution_to: Número de resolución final
            
        Returns:
            BytesIO: Archivo Excel en memoria
        """
        # Consultar obligaciones en el rango de resoluciones con join a Client y Tenant
        # Solo seleccionar columnas específicas para evitar errores con columnas faltantes
        statement = (
            select(Obligation.id, Obligation.amount, Obligation.currency, Obligation.due_date, 
                   Obligation.issue_date, Obligation.status, Obligation.type, Obligation.description,
                   Obligation.client_id, Obligation.process_id, Obligation.resolution_number,
                   Obligation.resolution_year, Obligation.resolution_date, Obligation.radicado_number,
                   Obligation.resolution_assigned_at, Obligation.resolution_observations,
                   Client.identification, Client.name.label('client_name'), Client.address, 
                   Client.phone, Client.email, Client.city.label('client_city'), 
                   Client.department.label('client_department'),
                   Tenant.name.label('tenant_name'), Tenant.code.label('tenant_code'))
            .join(Client, Obligation.client_id == Client.id)
            .join(Tenant, Client.tenant_id == Tenant.id)
            .where(
                Obligation.resolution_number.isnot(None)
            )
        )
        
        # Obtener resultados
        results = db.exec(statement).all()
        
        # Filtrar resultados que cumplen con el rango de resolución
        filtered_results = []
        for row in results:
            try:
                # Extraer el número de resolución si tiene formato numérico
                if row.resolution_number:
                    resolution_numeric_part = ExcelExportService._extract_numeric_part(row.resolution_number)
                    
                    if resolution_numeric_part and resolution_from <= resolution_numeric_part <= resolution_to:
                        filtered_results.append(row)
            except:
                # Si hay error al procesar el número de resolución, lo ignoramos
                continue

        # Definir columnas según el formato de la Alcaldía de Girón
        column_names = [
            'N°',
            'Documento',
            'Nombre',
            'Dirección',
            'Departamento',
            'Ciudad',
            'Referencia'
        ]

        # Construir datos
        data = []
        for idx, row in enumerate(filtered_results, start=1):
            # Usar valores por defecto (para la Alcaldía de Girón)
            # Los campos de departamento y ciudad se completarán con valores por defecto
            departamento = 'Santander'
            ciudad = 'Girón'
            
            data.append({
                'N°': idx,
                'Documento': row.identification or '',
                'Nombre': row.client_name or '',
                'Dirección': row.address or '',
                'Departamento': departamento,
                'Ciudad': ciudad,
                'Referencia': row.resolution_number or ''
            })

        df = pd.DataFrame(data)

        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = 'Obligaciones'[:31]  # Limitar a 31 caracteres

        # Definir estilos
        header_font = Font(name='Arial', size=11, bold=True)
        title_font = Font(name='Arial', size=12, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='center')

        # Fila 1: Título principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_names))
        title_cell = ws['A1']
        title_cell.value = f'RELACIÓN DE ENVIOS FISICOS'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fila 2: Fecha de exportación
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(column_names))
        date_cell = ws['A2']
        date_cell.value = f'FECHA DE EXPORTACIÓN: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        date_cell.font = Font(name='Arial', size=10)
        date_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Fila 3: Rango de resoluciones
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(column_names))
        range_cell = ws['A3']
        range_cell.value = f'RANGO: {resolution_from} - {resolution_to}'
        range_cell.font = Font(name='Arial', size=10)
        range_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Fila 4: Encabezados
        for col_idx, column_title in enumerate(column_names, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = column_title
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Filas de datos (desde la fila 5 en adelante)
        for row_idx, row_data in enumerate(data, start=5):
            for col_idx, (col_name, cell_value) in enumerate(zip(column_names, row_data.values()), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.border = border
                cell.alignment = alignment

        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # N°
            'B': 12,  # Documento
            'C': 35,  # Nombre
            'D': 40,  # Dirección
            'E': 15,  # Departamento
            'F': 15,  # Ciudad
            'G': 15   # Referencia
        }
        
        for col_letter, width in column_widths.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def _extract_numeric_part(resolution_string: str) -> int:
        """
        Extrae la parte numérica de un número de resolución.
        Este método puede necesitar ajustes según el formato real de los números de resolución.
        """
        import re
        if not resolution_string:
            return 0
        numbers = re.findall(r'\d+', resolution_string)
        if numbers:
            # Tomar el último número encontrado en la cadena
            return int(numbers[-1])
        return 0
